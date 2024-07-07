import os
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill

# Configuration
mode = os.getenv('PROJECT_MODE')
if mode == 'prod':
    print("Testing is not allowed in production mode. Please chage Project Mode key to:'dev' and reopen terminal to run test.")
    exit()
image_folder = './aadhar.v4i.tensorflow/test'
upload_api_endpoint = 'https://api.imgbb.com/1/upload'
ocr_api_endpoint = 'http://192.168.1.12:3100/ocrAdhaar'
excel_file = 'Test_ResultAdhaar.xlsx'


upload_params = {
    'key': '898b7b839ae4b0dbc0af3d392816c73e',
    'expiration': 600
}
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

# Helper function to convert pixels to column width
def pixels_to_column_width(pixels):
    return pixels * 0.14

# Step 1: Create an Excel file with three columns
wb = Workbook()
ws = wb.active
ws.title = "Image Data"
ws.append(["images", "image_urls", "ocr_results"])

# Set column widths
ws.column_dimensions['A'].width = pixels_to_column_width(500)
ws.column_dimensions['B'].width = pixels_to_column_width(500)
ws.column_dimensions['C'].width = pixels_to_column_width(500)
wb.save(excel_file)

# Get total number of images
image_files = [f for f in os.listdir(image_folder) if os.path.isfile(os.path.join(image_folder, f))]
total_images = len(image_files)
print(f"There are {total_images} images in your test folder.")

# Step 2: Upload images and store URLs
print("Uploading started...")
uploaded_count = 0
maxUpload = 1
for image_name in image_files:
    if uploaded_count >= maxUpload:
        break
    image_path = os.path.join(image_folder, image_name)
    try:
        # Upload image with params
        with open(image_path, 'rb') as img_file:
            response = requests.post(upload_api_endpoint, params=upload_params, files={'image': img_file})
        
        if response.status_code == 200:
            image_url = response.json().get('data', {}).get('display_url', 'No URL returned')
        else:
            print(f"Error uploading {image_name}: {response.status_code} - {response.text}")
            image_url = 'Error during generating image_url'
    except Exception as e:
        print(f"Error uploading {image_name}: {str(e)}")
        # Handle the error or set a default value
        image_url = f'Error during generating image_url: {str(e)}'

    # Save to Excel
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # Insert image into the Excel sheet
    img = Image(image_path)
    img.height = 500
    img.width = 500
    cell_location = f'A{ws.max_row + 1}'
    ws.add_image(img, cell_location)
    ws.row_dimensions[ws.max_row + 1].height = 375  # 500 px high (approximately)

    # Append the image name and URL
    ws.append([image_name, image_url, ''])
    wb.save(excel_file)

    uploaded_count += 1
    print(f"{uploaded_count}/{total_images} images uploaded.")

print("All images uploaded successfully.")
wb.close()

# Step 3: Read URLs from Excel and get OCR results
wb = load_workbook(excel_file)
ws = wb.active
total_urls = ws.max_row - 1
print(f"OCR generation started for {total_urls} URLs.")
ocr_count = 0

for row in range(2, ws.max_row + 1):
    # Save to Excel
    wb = load_workbook(excel_file)
    ws = wb.active
    image_url = ws.cell(row=row, column=2).value
    if image_url and 'Error' in image_url:
        # Apply red fill to image_urls cell
        ws.cell(row=row, column=2).fill = red_fill
        continue
    
    if image_url and 'Error' not in image_url:
        try:
            response = requests.post(ocr_api_endpoint, json={'imgUrl': image_url})
            if response.status_code == 200:
                ocr_result = response.json()
            else:
                print(f"Error generating OCR for {image_url}: {response.status_code} - {response.text}")
                # Handle the error or set a default value
                ocr_result = response.json()
                ws.cell(row=row, column=3).fill = red_fill
        except Exception as e:
            print(f"Error generating OCR for {image_url}: {str(e)}")
            # Handle the error or set a default value
            ocr_result = f'Error during generating ocr_result: {str(e)}'
            ws.cell(row=row, column=3).fill = red_fill
    else:
        print(f"Skipping URL: {image_url}")
        # Handle the error or set a default value
        ocr_result = 'Error during generating ocr_result'
        ws.cell(row=row, column=3).fill = red_fill
        
    
    # Update OCR result in Excel
    ws.cell(row=row, column=3, value=str(ocr_result))
    wb.save(excel_file)

    ocr_count += 1
    print(f"{ocr_count}/{total_urls} OCR results generated.")

print(f"OCR results generated successfully and saved to {excel_file}.")
wb.close()

