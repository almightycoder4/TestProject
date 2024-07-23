import os
import boto3
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill

# Configuration
mode = os.getenv('PROJECT_MODE')
if mode == 'prod':
    print("Testing is not allowed in production mode. Please change Project Mode key to:'dev' and reopen terminal to run test.")
    exit()

image_folder = '../../../TestPan2'
ocr_api_endpoint = 'https://pawanmau01-testapi.hf.space/ocrPan'
excel_file = 'Test_ResultPan.xlsx'
red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
max_upload = 100
# S3 Configuration
s3_bucket_name = 'iduploadbucket'
s3 = boto3.client('s3')

def pixels_to_column_width(pixels):
    return pixels * 0.14

def initialize_excel(file_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Image Data"
    ws.append(["images", "image_urls", "ocr_results"])
    ws.column_dimensions['A'].width = pixels_to_column_width(500)
    ws.column_dimensions['B'].width = pixels_to_column_width(500)
    ws.column_dimensions['C'].width = pixels_to_column_width(500)
    wb.save(file_path)

def upload_image_to_s3(image_path, bucket_name):
    try:
        file_name = os.path.basename(image_path)
        s3.upload_file(image_path, bucket_name, file_name)
        image_url = f"https://{bucket_name}.s3.amazonaws.com/{file_name}"
        return image_url
    except Exception as e:
        return f'Error: {str(e)}'

def insert_image_to_excel(ws, image_path, image_url, row):
    img = Image(image_path)
    img.height = 500
    img.width = 500
    cell_location = f'A{row}'
    ws.add_image(img, cell_location)
    ws.row_dimensions[row].height = 375
    ws.append([os.path.basename(image_path), image_url, ''])

def get_ocr_result(image_url):
    try:
        response = requests.post(ocr_api_endpoint, json={'imgUrl': image_url})
        print(response.text, response.status_code, "responce from ocr api")
        if response.status_code == 200:
            return response.json()
        else:
            return f'Error: {response.status_code} - {response.text}'
    except Exception as e:
        return f'Error: {str(e)}'

def main():
    initialize_excel(excel_file)
    
    image_files = [f for f in os.listdir(image_folder) if os.path.isfile(os.path.join(image_folder, f)) and f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff'))]
    total_images = len(image_files)
    print(f"There are {total_images} images in your test folder.")
    
    uploaded_count = 0
    wb = load_workbook(excel_file)
    ws = wb.active
    
    print("Uploading started...")
    for image_name in image_files:
        if uploaded_count >= max_upload:
            break
        image_path = os.path.join(image_folder, image_name)
        image_url = upload_image_to_s3(image_path, s3_bucket_name)
        insert_image_to_excel(ws, image_path, image_url, ws.max_row + 1)
        wb.save(excel_file)
        uploaded_count += 1
        print(f"{uploaded_count}/{total_images} images uploaded.")
    
    print("All images uploaded successfully.")
    wb.close()

    wb = load_workbook(excel_file)
    ws = wb.active
    image_urls = [ws.cell(row=row, column=2).value for row in range(2, ws.max_row + 1)]
    
    print(f"OCR generation started for {len(image_urls)} URLs.")
    work_result = []
    
    for image_url in image_urls:
        ocr_result = get_ocr_result(image_url)
        work_result.append(ocr_result)
        print(f"Processed {len(work_result)}/{len(image_urls)} URLs")

    for row, ocr_result in enumerate(work_result, start=2):
        ws.cell(row=row, column=3).value = str(ocr_result)
        if 'Error' in ocr_result:
            ws.cell(row=row, column=3).fill = red_fill
    
    wb.save(excel_file)
    wb.close()

    print(f"OCR results generated successfully and saved to {excel_file}.")

if __name__ == "__main__":
    main()
